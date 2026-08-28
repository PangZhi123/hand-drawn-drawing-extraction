from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domain.models import ExtractedItem


HEADERS = ["文本", "置信度", "状态", "源文件", "页码", "X", "Y", "宽度", "高度"]


def build_workbook(items: list[ExtractedItem], output: Path, title: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "提取结果"
    sheet.append(HEADERS)
    for item in items:
        box = item.bbox
        sheet.append([
            item.text,
            item.confidence,
            "待复核" if item.needs_review else "已识别",
            item.source_file,
            item.page_number,
            box.x if box else None,
            box.y if box else None,
            box.width if box else None,
            box.height if box else None,
        ])
    review = workbook.create_sheet("待复核")
    review.append(HEADERS)
    for item in items:
        if item.needs_review:
            box = item.bbox
            review.append([item.text, item.confidence, "待复核", item.source_file, item.page_number,
                           box.x if box else None, box.y if box else None,
                           box.width if box else None, box.height if box else None])
    summary = workbook.create_sheet("分析摘要", 0)
    summary.append(["项目", "内容"])
    summary.append(["工作簿名称", title])
    summary.append(["提取项数", len(items)])
    summary.append(["待复核项数", sum(item.needs_review for item in items)])
    for current in workbook.worksheets:
        current.freeze_panes = "A2"
        current.auto_filter.ref = current.dimensions
        for cell in current[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for column in current.columns:
            width = min(50, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            current.column_dimensions[get_column_letter(column[0].column)].width = width
    workbook.save(output)


def build_analysis_workbook(data: dict, output: Path, title: str) -> None:
    """固定工作表结构，将模型返回的全部信息展平写入，同时保留完整JSON。"""
    import json

    workbook = Workbook()
    summary = workbook.active
    summary.title = "分析摘要"
    summary.append(["项目", "内容"])
    summary.append(["工作簿名称", title])
    summary.append(["处理页数", len(data.get("page_results", []))])
    review = data.get("automatic_review", {})
    summary.append(["自动复核", "已完成" if review.get("completed") else "未完成"])
    summary.append(["复核问题数", review.get("issue_count", 0)])

    fixed = workbook.create_sheet("固定关键数据")
    fixed.append(["字段", "值", "单位", "置信度", "源文件", "页码", "来源位置", "复核结果", "复核说明", "备注"])
    for item in data.get("fixed_key_data", []):
        fixed.append([
            item.get("field_name", ""), item.get("field_value", ""), item.get("unit", ""),
            item.get("confidence", ""), item.get("source_file", ""), item.get("page_number", ""),
            item.get("source_location", ""), item.get("review_status", ""),
            item.get("review_notes", ""), item.get("notes", ""),
        ])

    process = workbook.create_sheet("过程与曲线分析")
    process.append(["分析类别", "来源文件", "页码", "层级路径", "内容"])
    for category, records in data.get("process_analysis", {}).items():
        for index, record in enumerate(records if isinstance(records, list) else [records], 1):
            source = record.get("source", {}) if isinstance(record, dict) else {}
            content = record.get("content", record) if isinstance(record, dict) else record
            for path, value in _flatten(content):
                process.append([category, source.get("source_file", ""), source.get("page_number", ""),
                                f"{index}.{path}".rstrip("."), _cell_value(value)])

    uncertain = workbook.create_sheet("不确定与复核")
    uncertain.append(["类别", "字段/原文", "可能值", "原因", "详细信息"])
    for item in data.get("uncertain_items", []):
        uncertain.append(["模型不确定项", item.get("text", "") if isinstance(item, dict) else str(item),
                          item.get("possible_value", "") if isinstance(item, dict) else "",
                          item.get("reason", "") if isinstance(item, dict) else "", _cell_value(item)])
    for issue in review.get("issues", []):
        uncertain.append(["后端自动复核", issue.get("field", ""), issue.get("value", ""),
                          "；".join(issue.get("reasons", [])), _cell_value(issue)])

    observations = workbook.create_sheet("补充观察")
    observations.append(["序号", "内容"])
    for index, item in enumerate(data.get("additional_observations", []), 1):
        observations.append([index, _cell_value(item)])

    raw = workbook.create_sheet("完整结果JSON")
    raw.append(["序号", "JSON内容"])
    json_text = json.dumps(data, ensure_ascii=False, indent=2)
    for index, chunk_start in enumerate(range(0, len(json_text), 30000), 1):
        raw.append([index, json_text[chunk_start:chunk_start + 30000]])

    for current in workbook.worksheets:
        current.freeze_panes = "A2"
        current.auto_filter.ref = current.dimensions
        for cell in current[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in current.columns:
            width = min(60, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            current.column_dimensions[get_column_letter(column[0].column)].width = width
        for row in current.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(output)


def _flatten(value, prefix=""):
    if isinstance(value, dict):
        for key, child in value.items():
            yield from _flatten(child, f"{prefix}.{key}".strip("."))
    elif isinstance(value, list):
        for index, child in enumerate(value, 1):
            yield from _flatten(child, f"{prefix}[{index}]")
    else:
        yield prefix, value


def _cell_value(value):
    import json
    return json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value
