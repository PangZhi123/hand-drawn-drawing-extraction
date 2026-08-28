from openpyxl import load_workbook

from app.domain.models import BoundingBox, ExtractedItem
from app.services.excel import build_workbook


def test_workbook_contains_traceability_and_review_sheet(tmp_path):
    target = tmp_path / "result.xlsx"
    items = [
        ExtractedItem("桩号 K1", 0.96, "drawing.png", 1, BoundingBox(10, 20, 80, 24), False),
        ExtractedItem("尺寸?", 0.42, "drawing.png", 1, BoundingBox(12, 60, 70, 22), True),
    ]
    build_workbook(items, target, "测试工作簿")

    workbook = load_workbook(target)
    assert workbook.sheetnames == ["分析摘要", "提取结果", "待复核"]
    assert workbook["提取结果"].max_row == 3
    assert workbook["待复核"].max_row == 2
    assert workbook["待复核"]["D2"].value == "drawing.png"
